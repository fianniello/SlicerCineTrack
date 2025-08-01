import slicer
# from slicer import ScriptedLoadableModuleLogic
from slicer.ScriptedLoadableModule import *
import qt, vtk, ctk

import os, csv, re
import SimpleITK as sitk
import sitkUtils

class TrackLogic(ScriptedLoadableModuleLogic):
  """This class should implement all the actual
  computation done by your module.  The interface
  should be such that other python code can import
  this class and make use of the functionality without
  requiring an instance of the Widget.
  Uses ScriptedLoadableModuleLogic base class, available at:
  https://github.com/Slicer/Slicer/blob/master/Base/Python/slicer/ScriptedLoadableModule.py
  """

  def __init__(self):
    """
    Called when the logic class is instantiated. Can be used for initializing member variables.
    """
    slicer.app.pythonConsole().clear()
    ScriptedLoadableModuleLogic.__init__(self)
    self.timer = qt.QTimer()
    self.redBackground = None 
    self.greenBackground = None
    self.yellowBackground = None
    self.backgrounds = {
      "Red": self.redBackground,
      "Green": self.greenBackground,
      "Yellow": self.yellowBackground
    }

  def setDefaultParameters(self, customParameterNode):
    """
    Initialize parameter node with default settings.
    """
    customParameterNode.totalImages = 0
    customParameterNode.fps = 5.0  # frames (i.e. images) per second
    customParameterNode.opacity = 1.0  # 100 %
    customParameterNode.overlayAsOutline = True
    customParameterNode.overlayColor = [0, 0.7, 0]

  def loadImagesIntoSequenceNode(self, shNode, paths):
    """
    Loads the cine images located in the provided paths into 3D Slicer. They are
    placed within a sequence node and the loaded image nodes are deleted thereafter.
    :param shNode: node representing the subject hierarchy
    :param paths: list of paths to the 2D images to be imported
    """
    # NOTE: This represents a node within the MRML scene, not within the subject hierarchy
    imagesSequenceNode = None

    # Find all the image file names within the provided paths
    imageFiles = []
    # Only accept valid file formats
    fileFormats = ['.*\.mha', '.*\.dcm', '.*\.nrrd', '.*\.nii', '.*\.hdr','.*\.nhdr', '.*\.mhd']  
    for path in paths:
      validFormat = any(re.match(format, path) for format in fileFormats)
      if validFormat:
        imageFiles.append(path)
    imageFiles.sort()

    # We only want to create a sequence node if image files were found within the provided paths
    if len(imageFiles) != 0:
      imagesSequenceNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLSequenceNode",
                                                              "Image Nodes Sequence")

      # Create a progress/loading bar to display the progress of the images loading process
      progressDialog = qt.QProgressDialog("Loading cine images", "Cancel",
                                          0, len(imageFiles))
      progressDialog.minimumDuration = 0

      for fileIndex in range(len(imageFiles)):
        # If the 'Cancel' button was pressed, we want to return to a default state
        if progressDialog.wasCanceled:
          # Remove sequence node
          slicer.mrmlScene.RemoveNode(imagesSequenceNode)
          return None, True

        filepath = imageFiles[fileIndex]
        nodeName = (f"Image {fileIndex + 1} ({os.path.basename(filepath)})")

        loadedImageNode = slicer.util.loadVolume(filepath, {"singleFile": True, "show": False})
        loadedImageNode.SetName(nodeName)
        # Place image node into sequence
        imagesSequenceNode.SetDataNodeAtValue(loadedImageNode, str(fileIndex))
        # Remove loaded image node
        imageID = shNode.GetItemByDataNode(loadedImageNode)
        shNode.RemoveItem(imageID)

        #  Update how far we are in the progress bar
        progressDialog.setValue(fileIndex + 1)

        # This render step is needed for the progress bar to visually update in the GUI
        slicer.util.forceRenderAllViews()
        slicer.app.processEvents()

      print(f"{len(imageFiles)} cine images were loaded into 3D Slicer")

      # We do the following to clear the view of the slices. I expected {"show": False} to
      # prevent anything from being shown at all, but the first loaded image will appear in the
      # foreground. This seems to be a bug in 3D Slicer.
      self.clearSliceForegrounds()

    return imagesSequenceNode, False

  def getColumnNamesFromTransformsInput(self, filepath):
      
    fileName = os.path.basename(filepath)
    fileExtension = os.path.splitext(filepath)[1]

    if re.match('.*\.(csv|xls|xlsx|txt)', filepath):
      # Check that the transforms file is a .csv type
      if filepath.endswith('.csv'):
        encodings = ["utf-8-sig", "cp1252", "iso-8859-1", "latin1"]
        for encoding in encodings:
          try:
            with open(filepath, "r", encoding = encoding) as f:
              reader = csv.reader(f)
              headers = next(reader)
              # if we can read without error, break of the encoding loop
              break
          except:
            print(f"Encoding {encoding} failed, trying next encoding")
        return headers
   
      if filepath.endswith('.txt'):
        with open(filepath, "r") as f:
          headers = next(f).strip().split(',')
          return headers
      if filepath.endswith('.xlsx'):
        try:
          import openpyxl
        except ModuleNotFoundError:
          if slicer.util.confirmOkCancelDisplay(f"To load {fileName}, install the 'openpyxl' Python package. Click OK to install now."):
            try:
              # Create a loading popup
              messageBox = qt.QMessageBox()
              messageBox.setIcon(qt.QMessageBox.Information)
              messageBox.setWindowTitle("Package Installation")
              messageBox.setText("Installing 'openpyxl' package...")
              messageBox.setStandardButtons(qt.QMessageBox.NoButton)
              messageBox.show()
              slicer.app.processEvents()

              slicer.util.pip_install('openpyxl')
              openpyxl = __import__('openpyxl')

              messageBox.setText(f"'openpyxl' package installed successfully. {fileName} will now load.")
              slicer.app.processEvents()  # Process events to allow the dialog to update
              qt.QTimer.singleShot(3000, messageBox.accept)

              # Wait for user interaction
              while messageBox.isVisible():
                slicer.app.processEvents()
            except:
              slicer.util.warningDisplay(f"{fileName} file failed to load.\nPlease load a .csv or .txt file instead. ",
                                          "Failed to Load File")
              return
          else:
            slicer.util.warningDisplay(f"{fileName} failed to load.\nPlease load a .csv or .txt file instead. ",
                                      "Failed to Load File")
            return
        openpyxl = __import__('openpyxl')
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        headers = next(sheet.iter_rows(values_only=True))
        # print(headers)
        return headers
      elif filepath.endswith('.xls'):
        try:
          import xlrd
        except ModuleNotFoundError:
          if slicer.util.confirmOkCancelDisplay(f"To load {fileName}, install the 'xlrd' Python package. Click OK to install now."):
            try:
              # Create a loading popup
              messageBox = qt.QMessageBox()
              messageBox.setIcon(qt.QMessageBox.Information)
              messageBox.setWindowTitle("Package Installation")
              messageBox.setText("Installing 'xlrd' package...")
              messageBox.setStandardButtons(qt.QMessageBox.NoButton)
              messageBox.show()
              slicer.app.processEvents()

              slicer.util.pip_install('xlrd')
              xlrd = __import__('xlrd')


              messageBox.setText(f"'xlrd' package installed successfully. {fileName} will now load.")
              slicer.app.processEvents()  # Process events to allow the dialog to update
              qt.QTimer.singleShot(3000, messageBox.accept)

              # Wait for user interaction
              while messageBox.isVisible():
                slicer.app.processEvents()

              messageBox.hide()  # Hide the message box
            except:
              slicer.util.warningDisplay(f"{fileName} file not loaded.\nPlease load a .csv or .txt file instead. ",
                              "Failed to Load File")
              return
          else:
            slicer.util.warningDisplay(f"{fileName} file not loaded.\nPlease load a .csv or .txt file instead. ",
                                      "Failed to Load File")
            return 
        xlrd = __import__('xlrd')
        wb = xlrd.open_workbook(filepath)
        sheet = wb.sheet_by_index(0)
        return sheet.row_values(0)
    
    # if we get here, we failed to read the the headers -> print out warning and return a empty list for headers   
    slicer.util.warningDisplay(f"Cannot read header row from {fileName}.\nPlease load another file instead. ",
                                  "Failed to Load File")
    return []

  def validateTransformsInput(self, filepath, numImages,headers):
    """
    Checks to ensure that the data in the provided transformation file is valid and matches the
    number of 2D images that have been loaded into 3D Slicer.
    :param filepath: path to the transforms file (which should be a .csv file)
    :param numImages: the number of cine images that have already been loaded
    """
    # NOTE: The current logic of this function will only ensure that the first {numImages}
    # transformations found within the CSV file are valid, so playback can occur. The playback will
    # still occur if later transformations after the first {numImages} transformations are corrupt.
    transformationsList = []
    fileName = os.path.basename(filepath)
    fileExtension = os.path.splitext(filepath)[1]
    headerX = headers[0]
    headerY = headers[1]
    headerZ = headers[2]
    if re.match('.*\.(csv|xls|xlsx|txt)', filepath):
      # Check that the transforms file is a .csv type
      if filepath.endswith('.csv'):
        encodings = ["utf-8-sig", "cp1252", "iso-8859-1", "latin1"]
        for encoding in encodings:
          try:
            with open(filepath, "r", encoding = encoding) as f:
              # Using a DictReader allows us to recognize the CSV header
              reader = csv.DictReader(f)
              for row in reader:
                # Extract floating point values from row
                transformationsList.append([float(row[headerX]), float(row[headerY]), float(row[headerZ])])
              
              # if we can read the file without error, break the encoding loop
              break
          except:
            print(f"Encoding {encoding} failed, trying next encoding")
          
      if len(transformationsList) == 0 and filepath.endswith('.csv'):
        slicer.util.warningDisplay(f"{fileName} file failed to load.\nPlease load another file instead. ",
                                  "Failed to Load File")
        return
      
      # Check that the transforms file is a .txt type
      elif filepath.endswith('.txt'):
        with open(filepath, "r") as f:
          next(f)
          for line in f:
            values = line.strip().split(',')
            try:
              x, y, z = map(float, values)
              transformationsList.append([x, y, z])
            except:
              # If there was an error reading the values, break out because we can't/shouldn't
              # perform the playback if the transformation data is corrupt or missing.
              slicer.util.warningDisplay(f"An error was encountered while reading the {fileExtension} file: "
                                   f"{fileName}",
                                   "Validation Error")
              break

      # Check that the transforms file is a .xlsx type
      elif filepath.endswith('.xlsx') or filepath.endswith('.xlsx'):
        openpyxl = __import__('openpyxl')        
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        rows = iter(sheet.iter_rows(values_only=True))
        header_row = next(rows)
        
        x_index = header_row.index(headerX)
        y_index = header_row.index(headerY)
        z_index = header_row.index(headerZ)

        for row in rows:
          try:
            x, y, z = map(float, [row[x_index], row[y_index], row[z_index]])
            transformationsList.append([x,y,z])
          except Exception as e:
            print(e)
            slicer.util.warningDisplay(f"{fileName} file failed to load.\nPlease load a .csv or .txt file instead. ",
                                      "Failed to Load File")
            break
        
      # Check that the transforms file is a .xls type
      elif filepath.endswith('.xls'):
        xlrd = __import__('xlrd')    
        workbook = xlrd.open_workbook(filepath)
        sheet = workbook.sheet_by_index(0)
        header_row = sheet.row_values(0)
        
        x_index = header_row.index(headerX)
        y_index = header_row.index(headerY)
        z_index = header_row.index(headerZ)
        
        for row_idx in range(1, sheet.nrows):  # Start from the second row, assuming first row is header
          values = sheet.row_values(row_idx)
          try:
            x, y, z = map(float, [values[x_index], values[y_index], values[z_index]])
            transformationsList.append([x, y, z])
          except:
            # If there was an error reading the values, break out because we can't/shouldn't
            # perform the playback if the transformation data is corrupt or missing.
            slicer.util.warningDisplay(f"An error was encountered while reading the {fileExtension} file: "
                                     f"{fileName}",
                                     "Validation Error")
            break


      if len(transformationsList) == numImages:
        return transformationsList
      else:
        # Extension will not create transforms nodes if the number of cine images and
        # the number of rows in the transforms file are not equal
        print(os.path.basename(filepath))
        slicer.util.warningDisplay(f"Error loading transforms file. Ensure proper formatting and matching number of transforms to cine images",
                           "Validation Error")
        
        return None

  def createTransformNodesFromTransformData(self, shNode, transforms, numImages):
    """
    For every image and it's matching transformation, create a transform node which will hold
    the transformation data for that image wthin 3D Slicer. Place them in a sequence node.
    :param shNode: node representing the subject hierarchy
    :param transforms: list of transforms extrapolated from the transforms .csv file
    :param numImages: number of 2D images loaded into 3D Slicer
    """
    # NOTE: This represents a node within the MRML scene, not within the subject hierarchy
    transformsSequenceNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLSequenceNode",
                                                                "Transform Nodes Sequence")

    # Create a progress/loading bar to display the progress of the node creation process
    progressDialog = qt.QProgressDialog("Creating Transform Nodes From Transformation Data", "Cancel",
                                        0, numImages)
    progressDialog.minimumDuration = 0

    # 3D Slicer works with 4x4 transform matrices internally
    LPSToRASMatrix = vtk.vtkMatrix4x4()
    LPSToRASMatrix.SetElement(0, 0, -1)
    LPSToRASMatrix.SetElement(1, 1, -1)

    # NOTE: It is very important that we loop using the number of 2D images loaded, versus the size
    # of the transforms array/list. This is because we may provide a CSV with more transforms than
    # needed, but we only need to create as many transform nodes as there are 2D images.
    for i in range(numImages):
      # If the 'Cancel' button was pressed, we want to return to a default state
      if progressDialog.wasCanceled:
        # Remove sequence node
        shNode.RemoveNode(transformsSequenceNode)
        return None

      # 3D Slicer uses the RAS (Right, Anterior, Superior) basis for their coordinate system.
      # However, the transformation data we use was generated outside of 3D Slicer, using DICOM
      # images, which corresponds to the LPS (Left, Prosterier, Superior) basis. In order to use
      # this data, we must convert it from LPS to RAS, in order to correctly transform the images
      # we load into 3D Slicer. See the following links for more detail:
      # https://www.slicer.org/wiki/Coordinate_systems#Anatomical_coordinate_system
      # https://github.com/Slicer/Slicer/blob/main/Libs/MRML/Core/vtkITKTransformConverter.h#L246
      # This is a simple conversion. It can be mathematically represented as:
      # /ΔLR\   /-1  0  0  0\   /X\
      # |ΔPA| = | 0 -1  0  0| * |Y|
      # |ΔIS|   | 0  0  1  0|   |Z|
      # \ 0 /   \ 0  0  0  1/   \0/
      # Where X, Y, and Z represent the transformation in LPS.

      # Convert transform from LPS to RAS
      currentTransform = transforms[i]
      currentTransform.append(0) # Needs to be 4x1 to multiply with a 4x4
      convertedTransform = [0, 0, 0, 0]
      LPSToRASMatrix.MultiplyPoint(currentTransform, convertedTransform)

      # Create a transform matrix from the converted transform
      transformMatrix = vtk.vtkMatrix4x4()
      transformMatrix.SetElement(0, 3, convertedTransform[0])  # LR translation
      transformMatrix.SetElement(1, 3, convertedTransform[1])  # PA translation
      transformMatrix.SetElement(2, 3, convertedTransform[2])  # IS translation

      # Create a LinearTransform node to hold our transform matrix
      transformNode = \
             slicer.mrmlScene.AddNewNodeByClass("vtkMRMLLinearTransformNode", f"Transform {i + 1}")
      transformNode.ApplyTransformMatrix(transformMatrix)

      # Add the transform node to the transforms sequence node
      transformsSequenceNode.SetDataNodeAtValue(transformNode, str(i))
      # Remove the transform node
      transformNodeID = shNode.GetItemByDataNode(transformNode)
      shNode.RemoveItem(transformNodeID)

      # Update how far we are in the progress bar
      progressDialog.setValue(i + 1)

      # This render step is needed for the progress bar to visually update in the GUI
      slicer.util.forceRenderAllViews()
      slicer.app.processEvents()

    print(f"{numImages} transforms were loaded into 3D Slicer as transform nodes")
    return transformsSequenceNode

  def clearSliceForegrounds(self):
    """
    Clear each slice view from having anything visible in the foreground. This often happens
    inadvertently when using loadVolume() with "show" set to False.
    """
    layoutManager = slicer.app.layoutManager()
    for viewName in layoutManager.sliceViewNames():
      layoutManager.sliceWidget(viewName).mrmlSliceCompositeNode().SetForegroundVolumeID("None")

  def visualize(self, sequenceBrowser, sequenceNode2DImages, segmentationLabelMapID,
                    sequenceNodeTransforms, opacity, overlayAsOutline, overlayThickness, show=False, customParamNode=None):
    """
    Visualizes the image data (2D images and 3D segmentation overlay) within the slice views and
    enables the alignment of the 3D segmentation label map according to the transformation data.
    :param sequenceBrowser: sequence browser node used to control the playback operation
    :param sequenceNode2DImages: sequence node containing the 2D images
    :param segmentationLabelMapID: subject hierarchy ID of the 3D segmentation label map
    :param sequenceNodeTransforms: sequence node containing the transforms
    :param opacity: opacity value of overlay layer (3D segmentation label map layer)
    :param overlayAsOutline: whether to show the overlay as an outline or a filled region
    """
    shNode = slicer.mrmlScene.GetSubjectHierarchyNode()
    layoutManager = slicer.app.layoutManager()

    # The proxy image node represents the current selected image within the sequence
    proxy2DImageNode = sequenceBrowser.GetProxyNode(sequenceNode2DImages)
    # The proxy transform node represents the current selected transform within the sequence
    proxyTransformNode = sequenceBrowser.GetProxyNode(sequenceNodeTransforms)
    labelMapNode = shNode.GetItemDataNode(segmentationLabelMapID)

    displayNode = labelMapNode.GetDisplayNode()
    if displayNode:
      # Ensure the color node is properly set and updated
      colorNode = displayNode.GetColorNode()
      if colorNode:
        # Force the color node to be re-applied
        displayNode.SetAndObserveColorNodeID(colorNode.GetID())
        colorNode.Modified()
        
      # NOTE: Removed automatic override of label 1 color to prevent conflicts with user-selected colors
      # The color buttons should control all label colors, including label 1

      displayNode.SetSliceIntersectionThickness(overlayThickness)
      
      # Force the display node to update
      displayNode.Modified()

    if proxy2DImageNode.GetImageData().GetDataDimension() == 2:
      sliceWidget = self.getSliceWidget(layoutManager, proxy2DImageNode)

      name = None
      fitSlice = None
      if sliceWidget is not None:
        name = sliceWidget.sliceViewName
          
        sliceCompositeNode = sliceWidget.mrmlSliceCompositeNode()

        volumesLogic = slicer.modules.volumes.logic()
        
        # Checks if the current slice node is not showing an image
        fitSlice = False
        if sliceCompositeNode.GetLabelVolumeID() is None:
          fitSlice = True
        
        sliceCompositeNode.SetLabelVolumeID(labelMapNode.GetID())
        sliceCompositeNode.SetLabelOpacity(opacity)
        
        sliceCompositeNode.SetBackgroundVolumeID(labelMapNode.GetID())
        
        # Get the current slice node
        sliceNode = sliceWidget.mrmlSliceNode()

        # Display the label map overlay as an outline
        sliceNode.SetUseLabelOutline(overlayAsOutline)

        # Set the background volume for the current slice view
        sliceCompositeNode.SetBackgroundVolumeID(proxy2DImageNode.GetID())

        # Translate the 3D segmentation label map using the transform data
        if proxyTransformNode is not None:
          labelMapNode.SetAndObserveTransformNodeID(proxyTransformNode.GetID())
        
        sliceNode.SetSliceVisible(True)

      # Make the 3D segmentation visible in the 3D view
      tmpIdList = vtk.vtkIdList() # The nodes you want to display need to be in a vtkIdList
      tmpIdList.InsertNextId(segmentationLabelMapID)
      threeDViewNode = layoutManager.activeMRMLThreeDViewNode()
      shNode.ShowItemsInView(tmpIdList, threeDViewNode)
      
      # Ensure 3D viewer properly reflects any color table changes
      if threeDViewNode:
        # Force update any volume rendering display nodes for the 3D view
        labelMapNode = shNode.GetItemDataNode(segmentationLabelMapID)
        if labelMapNode:
          # Update the main display node
          displayNode = labelMapNode.GetDisplayNode()
          if displayNode:
            displayNode.Modified()
            
          # CRITICAL: Force volume rendering to update colors
          volumeRenderingLogic = slicer.modules.volumerendering.logic()
          volumeRenderingDisplayNode = volumeRenderingLogic.GetFirstVolumeRenderingDisplayNode(labelMapNode)
          
          if volumeRenderingDisplayNode:
            # Force volume rendering to refresh with new color table
            volumeRenderingDisplayNode.Modified()
            volumePropertyNode = volumeRenderingDisplayNode.GetVolumePropertyNode()
            if volumePropertyNode:
              volumePropertyNode.Modified()
            
            # Force visibility update to trigger refresh
            wasVisible = volumeRenderingDisplayNode.GetVisibility()
            volumeRenderingDisplayNode.SetVisibility(False)
            slicer.app.processEvents()
            volumeRenderingDisplayNode.SetVisibility(wasVisible)
            
          # Update all display nodes including volume rendering
          for displayNodeIndex in range(labelMapNode.GetNumberOfDisplayNodes()):
            volumeDisplayNode = labelMapNode.GetNthDisplayNode(displayNodeIndex)
            if volumeDisplayNode:
              volumeDisplayNode.Modified()
              if volumeDisplayNode.IsA("vtkMRMLVolumeRenderingDisplayNode"):
                # Update volume rendering to reflect color changes
                volumeProperty = volumeDisplayNode.GetVolumePropertyNode()
                if volumeProperty:
                  volumeProperty.Modified()
          
          # Force the label map node itself to update
          labelMapNode.Modified()
          
          # Force 3D view to re-render
          layoutManager = slicer.app.layoutManager()
          if layoutManager:
            for threeDViewIndex in range(layoutManager.threeDViewCount):
              threeDWidget = layoutManager.threeDWidget(threeDViewIndex)
              if threeDWidget and threeDWidget.threeDView():
                threeDWidget.threeDView().forceRender()

      # If the sliceNode is now showing an image, fit the slice view to the current background image   
      if fitSlice:
        sliceWidget.fitSliceToBackground()
      
      # Preserve previous slices
      # If a background node for the specified orientation exists, update it with the current slice
      # Otherwise, create a new background node and set it as the background for the specified orientation
      
      if name in self.backgrounds:
        background = getattr(self, name.lower() + 'Background')
        if background is None:
          # Create a new background node for the orientation
          setattr(self, name.lower() + 'Background', volumesLogic.CloneVolume(slicer.mrmlScene,
                  proxy2DImageNode, f"{proxy2DImageNode.GetAttribute('Sequences.BaseName')}"))
        else:
          # Background exists, just replace the data to represent the next image in the sequence
          background.SetAndObserveImageData(proxy2DImageNode.GetImageData())
          background.SetAttribute("Sequences.BaseName", proxy2DImageNode.GetAttribute("Sequences.BaseName"))
      
      # Add the image name to the slice view background variable
      if name is not None:
        currentSlice = getattr(self, name.lower() + 'Background')
        currentSlice.SetName(proxy2DImageNode.GetAttribute('Sequences.BaseName'))

      # Set the background volumes for each orientation, if they exist
      for color in self.backgrounds:
        sliceViewWindow = slicer.app.layoutManager().sliceWidget(color).sliceView()
        sliceViewWindow.cornerAnnotation().RemoveAllObservers()
        currentSlice = getattr(self, color.lower() + 'Background')
        sliceViewWindow.cornerAnnotation().ClearAllTexts()
        # Add desired text to slice views that have a background node
        if currentSlice is not None:
          slicer.mrmlScene.GetNodeByID(f"vtkMRMLSliceCompositeNode{color}").SetBackgroundVolumeID(currentSlice.GetID())
          imageFile = slicer.mrmlScene.GetNodeByID(f"vtkMRMLSliceCompositeNode{color}").GetNodeReference('backgroundVolume') is not None
          if imageFile:
            imageFileNameText = slicer.mrmlScene.GetNodeByID(f"vtkMRMLSliceCompositeNode{color}").GetNodeReference('backgroundVolume').GetAttribute('Sequences.BaseName')
            # Place "Current Alignment" text in the slice view corner
            sliceViewWindow = slicer.app.layoutManager().sliceWidget(color).sliceView()
            sliceViewWindow.cornerAnnotation().SetText(0, imageFileNameText)
      
      for color in self.backgrounds:
        sliceViewWindow = slicer.app.layoutManager().sliceWidget(color).sliceView()
        if sliceViewWindow.cornerAnnotation().HasObserver(vtk.vtkCommand.ModifiedEvent):
          sliceViewWindow.cornerAnnotation().RemoveAllObservers()
        if sliceViewWindow.cornerAnnotation().HasObserver(vtk.vtkCornerAnnotation.UpperLeft):
          sliceViewWindow.cornerAnnotation().RemoveAllObservers()

      # Ensure "Current Alignment" text is displayed in the slice view corner only when required
      if show:
        if sliceWidget is not None:
          sliceView = sliceWidget.sliceView()
          sliceView.cornerAnnotation().SetText(vtk.vtkCornerAnnotation.UpperLeft, "Current Alignment")
      # Enable alignment of the 3D segmentation label map according to the transform data so that
      # the 3D segmentation label map overlays upon the ROI of the 2D images
      if proxyTransformNode is not None:
        labelMapNode.SetAndObserveTransformNodeID(proxyTransformNode.GetID())

      # Render changes
      # Force display node to update first
      displayNode = labelMapNode.GetDisplayNode()
      if displayNode:
        displayNode.Modified()
      labelMapNode.Modified()
      
      slicer.util.forceRenderAllViews()
      slicer.app.processEvents()

    else:
      sliceWidgets = self.getSliceWidgets(layoutManager, proxy2DImageNode)
      for sliceWidget in sliceWidgets: 
        name = None
        fitSlice = None
        if sliceWidget is not None:
          name = sliceWidget.sliceViewName
            
          sliceCompositeNode = sliceWidget.mrmlSliceCompositeNode()

          volumesLogic = slicer.modules.volumes.logic()
          
          # Checks if the current slice node is not showing an image
          fitSlice = False
          if sliceCompositeNode.GetLabelVolumeID() is None:
            fitSlice = True
          
          sliceCompositeNode.SetLabelVolumeID(labelMapNode.GetID())
          sliceCompositeNode.SetLabelOpacity(opacity)
          
          sliceCompositeNode.SetBackgroundVolumeID(labelMapNode.GetID())
          
          # Get the current slice node
          sliceNode = sliceWidget.mrmlSliceNode()

          # Display the label map overlay as an outline
          sliceNode.SetUseLabelOutline(overlayAsOutline)

          # Set the background volume for the current slice view
          sliceCompositeNode.SetBackgroundVolumeID(proxy2DImageNode.GetID())

          # Translate the 3D segmentation label map using the transform data
          if proxyTransformNode is not None:
            labelMapNode.SetAndObserveTransformNodeID(proxyTransformNode.GetID())
          
          sliceNode.SetSliceVisible(True)

        # Make the 3D segmentation visible in the 3D view
        tmpIdList = vtk.vtkIdList() # The nodes you want to display need to be in a vtkIdList
        tmpIdList.InsertNextId(segmentationLabelMapID)
        threeDViewNode = layoutManager.activeMRMLThreeDViewNode()
        shNode.ShowItemsInView(tmpIdList, threeDViewNode)

        # Ensure 3D viewer properly reflects any color table changes
        if threeDViewNode:
          # Force update any volume rendering display nodes for the 3D view
          labelMapNode = shNode.GetItemDataNode(segmentationLabelMapID)
          if labelMapNode:
            # Update the main display node
            displayNode = labelMapNode.GetDisplayNode()
            if displayNode:
              displayNode.Modified()
              
            # CRITICAL: Force volume rendering to update colors
            volumeRenderingLogic = slicer.modules.volumerendering.logic()
            volumeRenderingDisplayNode = volumeRenderingLogic.GetFirstVolumeRenderingDisplayNode(labelMapNode)
            
            if volumeRenderingDisplayNode:
              # Force volume rendering to refresh with new color table
              volumeRenderingDisplayNode.Modified()
              volumePropertyNode = volumeRenderingDisplayNode.GetVolumePropertyNode()
              if volumePropertyNode:
                volumePropertyNode.Modified()
              
              # Force visibility update to trigger refresh
              wasVisible = volumeRenderingDisplayNode.GetVisibility()
              volumeRenderingDisplayNode.SetVisibility(False)
              slicer.app.processEvents()
              volumeRenderingDisplayNode.SetVisibility(wasVisible)
              
            # Update all display nodes including volume rendering
            for displayNodeIndex in range(labelMapNode.GetNumberOfDisplayNodes()):
              volumeDisplayNode = labelMapNode.GetNthDisplayNode(displayNodeIndex)
              if volumeDisplayNode:
                volumeDisplayNode.Modified()
                if volumeDisplayNode.IsA("vtkMRMLVolumeRenderingDisplayNode"):
                  # Update volume rendering to reflect color changes
                  volumeProperty = volumeDisplayNode.GetVolumePropertyNode()
                  if volumeProperty:
                    volumeProperty.Modified()
            
            # Force the label map node itself to update
            labelMapNode.Modified()
            
            # Force 3D view to re-render
            layoutManager = slicer.app.layoutManager()
            if layoutManager:
              for threeDViewIndex in range(layoutManager.threeDViewCount):
                threeDWidget = layoutManager.threeDWidget(threeDViewIndex)
                if threeDWidget and threeDWidget.threeDView():
                  threeDWidget.threeDView().forceRender()

        # If the sliceNode is now showing an image, fit the slice view to the current background image   
        if fitSlice:
          sliceWidget.fitSliceToBackground()
        
        # Preserve previous slices
        # If a background node for the specified orientation exists, update it with the current slice
        # Otherwise, create a new background node and set it as the background for the specified orientation
        
        if name in self.backgrounds:
          background = getattr(self, name.lower() + 'Background')
          if background is None:
            # Create a new background node for the orientation
            setattr(self, name.lower() + 'Background', volumesLogic.CloneVolume(slicer.mrmlScene,
                    proxy2DImageNode, f"{proxy2DImageNode.GetAttribute('Sequences.BaseName')}"))
          else:
            # Background exists, just replace the data to represent the next image in the sequence
            background.SetAndObserveImageData(proxy2DImageNode.GetImageData())
            background.SetAttribute("Sequences.BaseName", proxy2DImageNode.GetAttribute("Sequences.BaseName"))
        
        # Add the image name to the slice view background variable
        if name is not None:
          currentSlice = getattr(self, name.lower() + 'Background')
          currentSlice.SetName(proxy2DImageNode.GetAttribute('Sequences.BaseName'))

        # Set the background volumes for each orientation, if they exist
        for color in self.backgrounds:
          sliceViewWindow = slicer.app.layoutManager().sliceWidget(color).sliceView()
          sliceViewWindow.cornerAnnotation().RemoveAllObservers()
          currentSlice = getattr(self, color.lower() + 'Background')
          sliceViewWindow.cornerAnnotation().ClearAllTexts()
          # Add desired text to slice views that have a background node
          if currentSlice is not None:
            slicer.mrmlScene.GetNodeByID(f"vtkMRMLSliceCompositeNode{color}").SetBackgroundVolumeID(currentSlice.GetID())
            imageFile = slicer.mrmlScene.GetNodeByID(f"vtkMRMLSliceCompositeNode{color}").GetNodeReference('backgroundVolume') is not None
            if imageFile:
              imageFileNameText = slicer.mrmlScene.GetNodeByID(f"vtkMRMLSliceCompositeNode{color}").GetNodeReference('backgroundVolume').GetAttribute('Sequences.BaseName')
              # Place "Current Alignment" text in the slice view corner
              sliceViewWindow = slicer.app.layoutManager().sliceWidget(color).sliceView()
              sliceViewWindow.cornerAnnotation().SetText(0, imageFileNameText)
        
        for color in self.backgrounds:
          sliceViewWindow = slicer.app.layoutManager().sliceWidget(color).sliceView()
          if sliceViewWindow.cornerAnnotation().HasObserver(vtk.vtkCommand.ModifiedEvent):
            sliceViewWindow.cornerAnnotation().RemoveAllObservers()
          if sliceViewWindow.cornerAnnotation().HasObserver(vtk.vtkCornerAnnotation.UpperLeft):
            sliceViewWindow.cornerAnnotation().RemoveAllObservers()

        # Enable alignment of the 3D segmentation label map according to the transform data so that
        # the 3D segmentation label map overlays upon the ROI of the 2D images
        if proxyTransformNode is not None:
          labelMapNode.SetAndObserveTransformNodeID(proxyTransformNode.GetID())

        # Render changes
        # Force display node to update first
        displayNode = labelMapNode.GetDisplayNode()
        if displayNode:
          displayNode.Modified()
        labelMapNode.Modified()
        
        slicer.util.forceRenderAllViews()
        slicer.app.processEvents()
  
  def getSliceWidget(self, layoutManager, imageNode):
    """
    This function helps to determine the slice widget that corresponds to the orientation of the
    provided image. (i.e. the slice widget that would display the image)
    :param layoutManager: node representing the MRML layout manager
    :param imageNode: node representing the 2D image
    """
    
    def get_anatomical_orientation(image):
      """
      Helper function for fixing images with incorrect anatomical orientation.
      Determine the anatomical orientation of an image based on its direction cosines.
      """
      direction = image.GetDirection()
      anatomical_labels = ['R', 'A', 'I', 'L', 'P', 'S']
      orientation = []

      # Determine the dominant direction for each axis
      for axis in range(3):
        # Extract the direction vector for the current axis
        vector = direction[axis::3]
        # Determine the index of the dominant direction
        max_index = max(range(3), key=lambda i: abs(vector[i]))
        # Append the corresponding anatomical label
        orientation.append(anatomical_labels[max_index + (3 if vector[max_index] < 0 else 0)])

      return ''.join(orientation)

    def reorient_image(image, orientation):
      """
      Helper function for fixing images oritentation.
      Reorient the image based on the anatomical orientation.
      """
      if image.GetSize()[0] == 1:
        # Single-slice in x-direction, check for sagittal, coronal, or axial
        if orientation[0] in ('L', 'R'):
          return sitk.DICOMOrient(image, 'PIR')  # Sagittal
        elif orientation[0] in ('A', 'P'):
          return sitk.DICOMOrient(image, 'LIA')  # Coronal
        else:
          return sitk.DICOMOrient(image, 'LPS')  # Axial
      elif image.GetSize()[1] == 1:
        # Single-slice in y-direction, check for sagittal, coronal, or axial
        if orientation[1] in ('L', 'R'):
          return sitk.DICOMOrient(image, 'PIR')  # Sagittal
        elif orientation[1] in ('A', 'P'):
          return sitk.DICOMOrient(image, 'LIA')  # Coronal
        else:
          return sitk.DICOMOrient(image, 'LPS')  # Axial
      return image
  
    # Determine the orientation of the image
    sitk_image = sitkUtils.PullVolumeFromSlicer(imageNode)
    if sitk_image.GetSize()[2] != 1:
      orientation = get_anatomical_orientation(sitk_image)
      reoriented_image = reorient_image(sitk_image, orientation)
      imageNode = sitkUtils.PushVolumeToSlicer(reoriented_image, None, name=imageNode.GetName())
    tmpMatrix = vtk.vtkMatrix4x4()
    if imageNode is not None:
      imageNode.GetIJKToRASMatrix(tmpMatrix)
      scanOrder = imageNode.ComputeScanOrderFromIJKToRAS(tmpMatrix)

      if scanOrder == "LR" or scanOrder == "RL":
        imageOrientation = "Sagittal"
      elif scanOrder == "AP" or scanOrder == "PA":
        imageOrientation = "Coronal"
      elif scanOrder == "IS" or scanOrder == "SI":
        imageOrientation = "Axial"
      else:
        print(f"Error: Unexpected image scan order {scanOrder}.")
        exit(1)

      # Find the slice widget that has the same orientation as the image
      sliceWidget = None
      for name in layoutManager.sliceViewNames():
        if layoutManager.sliceWidget(name).sliceOrientation == imageOrientation:
          sliceWidget = layoutManager.sliceWidget(name)

      if not sliceWidget:
        print(f"Error: A slice with the {imageOrientation} orientation was not found.")
        exit(1)

      return sliceWidget

  def getSliceWidgets(self, layoutManager, imageNode):
    """
    This function helps to determine the slice widgets that corresponds to the orientation of the
    provided image. (i.e. the slice widget that would display the image)
    :param layoutManager: node representing the MRML layout manager
    :param imageNode: node representing the 3D image
    """
    sliceWidgets = []
    for name in layoutManager.sliceViewNames():
      if layoutManager.sliceWidget(name).sliceOrientation == "Axial" or layoutManager.sliceWidget(name).sliceOrientation == "Sagittal" or layoutManager.sliceWidget(name).sliceOrientation == "Coronal":
        sliceWidgets.append(layoutManager.sliceWidget(name))
      else:
        print(f"Error: A slice with the required orientations was not found.")
        exit(1)
    return sliceWidgets